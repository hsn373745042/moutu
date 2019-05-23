import requests
import os
from bs4 import BeautifulSoup
import re
import threading
import time
import openpyxl

# 第一个多线程用来构造四级url和爬取五级url(存放图片的url)并下载图片
class MyThread(threading.Thread):
    def __init__(self,threadID,num):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.num = num

    def run(self):
        print("开始线程：" + self.num)
        crawler1(self.num)
        print("退出线程：" + self.num)

def crawler1(num):
    global l2,l0
    while len(l2) > 0:
        url = l2.pop(0)
        headers = {}
        res = requests.get(url,headers=headers,timeout=60)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text,'html.parser')
            o = soup.find(class_='pagenavi').find_all('a')[-2].text
            try:
                title = soup.find(class_='main-image').find('img')['alt'].replace('?','')
                os.mkdir('d:\\python\\meizitu\\' + title)
            except:
                try:
                    title = soup.find(class_='main-image').find('img')['alt'].replace('：','')
                    os.mkdir('d:\\python\\meizitu\\' + title)
                except:
                    title = soup.find(class_='main-image').find('img')['alt']
                    os.mkdir('d:\\python\\meizitu\\'+title)
            for x in range(1,int(o)+1):
                if x == 1:
                    urls = url
                    headers = {}
                    res = requests.get(urls, headers=headers, timeout=60)
                    if res.status_code == 200:
                        soup = BeautifulSoup(res.text, 'html.parser')
                        ur = soup.find(class_='main-image').find('img')['src']
                        name = re.findall('.*?(\d.*)', ur)[0].replace('/', '_')
                        headers = {}
                        res = requests.get(ur, headers=headers, timeout=60)
                        with open('d:/python/meizitu/' + title + '/' + name, 'wb') as f:
                            f.write(res.content)
                    else:
                        # 将爬取失败的四级url写入excel
                        add_url2([urls])
                else:
                    urls = url+'/'+str(x)
                    headers = {}
                    res = requests.get(urls,headers=headers,timeout=60)
                    if res.status_code == 200:
                        soup = BeautifulSoup(res.text, 'html.parser')
                        ur = soup.find(class_='main-image').find('img')['src']
                        name = re.findall('.*?(\d.*)', ur)[0].replace('/', '_')
                        headers = {}
                        res = requests.get(ur, headers=headers,timeout=60)
                        with open('d:/python/meizitu/'+title+'/'+name,'wb') as f:
                            f.write(res.content)
                    else:
                        add_url2([urls])
                        pass
            print('%s号完成爬取一组图片 %s' % (num,time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        else:
            l0.append(url)
            print('%s号遇到无法爬取的三级url %s' % (num,time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))

# 第二个多线程用来爬取五级url(存放图片的url)并下载图片
class my_thread(threading.Thread):
    def __init__(self,num):
        threading.Thread.__init__(self)
        self.num = num

    def run(self):
        print("开始线程：" + self.num)
        crawler2(self.num)
        print("退出线程：" + self.num)

def crawler2(num):
    global l3,l4
    while len(l3) > 0:
        url = l3.pop(0)
        headers = {}
        res = requests.get(url, headers=headers, timeout=60)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            urls = soup.find(class_='main-image').find('img')['src']
            name = re.findall('.*?(\d.*)', urls)[0].replace('/', '_')
            try:
                title = soup.find(class_='main-image').find('img')['alt'].replace('?','')
            except:
                try:
                    title = soup.find(class_='main-image').find('img')['alt'].replace('：','')
                except:
                    title = soup.find(class_='main-image').find('img')['alt']
            headers = {}
            res = requests.get(urls,headers=headers,timeout=60)
            with open('d:/python/meizitu/'+title+'/'+name,'wb') as f:
                f.write(res.content)
            print('%s号完成爬取一张图片 %s' % (num,time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        else:
            l4.append([url])

# 读取爬取失败的三级url
def read_url1():
    l = []
    wb = openpyxl.load_workbook('./meizi.xlsx')
    sheet = wb['url1']
    for i in sheet['A']:
        l.append(i.value)
    return l

# 写入爬取失败的四级url
def add_url2(l):
    wb = openpyxl.load_workbook('./meizi.xlsx')
    sheet = wb['url2']
    sheet.append(l)

# 读取爬取失败的四级url
def read_url2():
    l = []
    wb = openpyxl.load_workbook('./meizi.xlsx')
    sheet = wb['url2']
    for i in sheet['A']:
        l.append(i.value)
    return l

# 创建excel并保存爬取失败的四级url
def my_url(s):
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 添加工作表
    sheet = wb.active
    # 给工作表命名
    sheet.title = 'url'
    for i in s:
        sheet.append(i)
    # 保存工作簿并命名
    wb.save('meizi1.xlsx')

if __name__ == '__main__':

    start = time.time()
    # 爬剩余首页、性感、日本、台湾、清纯妹子图片
    l2 = read_url1()  # 读取爬取失败的三级url
    l0 = []           # 用来放爬取失败的三级url

    # 创建15个新线程
    thread = {}
    for i in range(15):
        thread[str(i)] = MyThread(str(i), str(i))
    # 开启新线程
    for i in range(15):
        thread[str(i)].start()
    # 等待至线程中止
    for i in range(15):
        thread[str(i)].join()

    # 爬取之前爬取失败的三级url
    while len(l0) > 0:
        url = l0[0]
        headers = {}
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text,'html.parser')
            urls = soup.find(class_='main-image').find('img')['src']
            o = soup.find(class_='pagenavi').find_all('a')[-2].text
            try:
                title = soup.find(class_='main-image').find('img')['alt'].replace('?','')
                os.mkdir('d:\\python\\meizitu\\' + title)
            except:
                try:
                    title = soup.find(class_='main-image').find('img')['alt'].replace('：','')
                    os.mkdir('d:\\python\\meizitu\\' + title)
                except:
                    title = soup.find(class_='main-image').find('img')['alt']
                    os.mkdir('d:\\python\\meizitu\\'+title)
            for x in range(1,int(o)+1):
                if x == 1:
                    name = re.findall('.*?(\d.*)',urls)[0].replace('/', '_')
                    headers = {}
                    res = requests.get(urls, headers=headers,timeout=60)
                    with open('d:/python/meizitu/' + title + '/' + name, 'wb') as f:
                        f.write(res.content)
                else:
                    urls = url+'/'+str(x)
                    headers = {}
                    res = requests.get(urls,headers=headers,timeout=60)
                    if res.status_code == 200:
                        soup = BeautifulSoup(res.text, 'html.parser')
                        ur = soup.find(class_='main-image').find('img')['src']
                        name = re.findall('.*?(\d.*)', ur)[0].replace('/', '_')
                        headers = {}
                        res = requests.get(ur, headers=headers,timeout=60)
                        with open('d:/python/meizitu/'+title+'/'+name,'wb') as f:
                            f.write(res.content)
                    else:
                        add_url2([urls])
                        pass
            l0.pop(0)
            print('成功爬取到一组三级url内的图片')
        elif 400 <= res.status_code < 500:
            l0.pop(0)
            print('一组三级url无法爬取')
        else:
            pass

    l3 = list(set(read_url2()))  # 获取去重后的爬取失败的四级url
    l4 = []                      # 用来放爬取失败的四级url

    # 创建新线程
    thread = {}
    for i in range(15):
        thread[str(i)] = my_thread(str(i))
    # 开启新线程
    for i in range(15):
        thread[str(i)].start()
    # 等待至线程中止
    for i in range(15):
        thread[str(i)].join()
    # 保存爬取失败的四级url
    my_url(l4)
    print("退出主线程，耗时%s" % (time.time() - start))